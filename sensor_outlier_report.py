import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import sys
import requests
import json

# ==============================================================================
# [사용자 설정] 환경변수 확인
# ==============================================================================
POSCO_GPT_KEY = os.environ.get('POSCO_GPT_KEY')

# ==============================================================================

try:
    from sklearn.ensemble import IsolationForest
except ImportError:
    print("[Error] scikit-learn 설치가 필요합니다. (pip install scikit-learn requests)")
    sys.exit()

# ==========================================
# [핵심] 사내 AI API 연동 함수
# ==========================================
def call_corporate_ai_api(prompt_text):
    if not POSCO_GPT_KEY:
        return "[설정 오류] 환경변수 POSCO_GPT_KEY가 없습니다."

    try:
        url = "http://aigpt.posco.net/gpgpta01-gpt/gptApi/personalApi"

        token_value = POSCO_GPT_KEY
        if not token_value.startswith("Bearer "):
            token_value = f"Bearer {token_value}"

        headers = {
            "Authorization": token_value,
            "Content-Type": "application/json"
        }

        payload = {
            "model": "gpt-5.2",
            "messages": [
                {"role": "system", "content": "당신은 포스코 퓨처엠의 설비 예지보전 전문가입니다. 수치를 기반으로 냉철하게 설비 위험도를 판단하십시오."},
                {"role": "user", "content": prompt_text}
            ],
            "frequency_penalty": 0.0,
            "temperature": 0.3
        }

        response = requests.post(url, headers=headers, json=payload, timeout=60)
        response.encoding = 'utf-8'

        if response.status_code == 200:
            try:
                result = response.json()
                if 'choices' in result:
                    return result['choices'][0]['message']['content'].strip()
                elif 'response' in result:
                    return result['response']
                else:
                    return str(result)
            except json.JSONDecodeError:
                return response.text.strip()
        else:
            return f"[API Error] 상태코드: {response.status_code}"

    except Exception as e:
        return f"[시스템 오류] {str(e)}"


# ==========================================
# 1. 데이터 로드 (안정성 강화 버전)
# ==========================================
def load_data(path):
    print(f"[1/5] 데이터 로딩 중: {os.path.basename(path)}")
    encodings = ['cp949', 'euc-kr', 'utf-8']
    df_raw = None
    separators = [',', ';', '\t']

    for sep in separators:
        for enc in encodings:
            try:
                df_raw = pd.read_csv(path, sep=sep, header=None, encoding=enc, low_memory=False)
                if df_raw.shape[1] > 1:
                    break
                else:
                    df_raw = None
            except:
                continue
        if df_raw is not None:
            break

    if df_raw is None or len(df_raw) < 3:
        raise ValueError("파일을 읽을 수 없습니다. (형식이 올바르지 않음)")

    tags = df_raw.iloc[0].astype(str).str.strip()
    names = df_raw.iloc[1].astype(str).str.strip()

    time_col_idx = 0
    for i, val in enumerate(names):
        if 'time' in val.lower() or 'date' in val.lower():
            time_col_idx = i
            break

    sensor_map = {}
    valid_indices = []

    for i in range(len(df_raw.columns)):
        if i == time_col_idx:
            continue
        if not tags.iloc[i] or tags.iloc[i] == 'nan':
            continue
        sensor_map[i] = {'tag': tags.iloc[i], 'name': names.iloc[i]}
        valid_indices.append(i)

    df_data = df_raw.iloc[2:].copy()
    df_data.rename(columns={time_col_idx: 'Time'}, inplace=True)

    for i in valid_indices:
        df_data[i] = pd.to_numeric(df_data[i], errors='coerce')
        df_data[i] = df_data[i].interpolate(method='linear').bfill().fillna(0)

    return df_data, sensor_map


# ==========================================
# 2. 위험 등급 판정 함수
# ==========================================
def classify_risk(score, max_z, max_consecutive, count_ratio, change_ratio):
    """
    3단계 위험 등급 판정:
      위험 : Isolation Forest 심각도 > 0.15 OR Z-score > 5 OR 연속이상 > 20건 OR 이상비율 > 5%
      주의 : Isolation Forest 심각도 > 0.08 OR Z-score > 3 OR 연속이상 > 10건 OR 이상비율 > 2%
      정상 : 위 기준 미달
    """
    if score > 0.15 or max_z > 5 or max_consecutive > 20 or count_ratio > 5:
        return "위험"
    elif score > 0.08 or max_z > 3 or max_consecutive > 10 or count_ratio > 2:
        return "주의"
    else:
        return "정상"


# ==========================================
# 3-A. 단위 추출 헬퍼
# ==========================================
def extract_unit(sensor_name):
    """센서명에서 단위 문자열 추출 (예: [℃] → '℃')"""
    import re
    m = re.search(r'\[([^\]]+)\]', sensor_name)
    if m:
        inner = m.group(1)
        for u in ['℃', 'A', '%', 'kW', 'V', 'kPa', 'Pa', 'rpm', 'Hz']:
            if u in inner:
                return u
    # 키워드 fallback
    upper = sensor_name.upper()
    if '온도' in sensor_name or 'TEMP' in upper:
        return '℃'
    if '전류' in sensor_name or 'CURR' in upper:
        return 'A'
    if '개도' in sensor_name or '출력' in sensor_name or 'VALVE' in upper:
        return '%'
    if '전력' in sensor_name or 'POWER' in upper:
        return 'kW'
    if '전압' in sensor_name:
        return 'V'
    return ''


# ==========================================
# 3-B. 자연어 문제 요약 생성
# ==========================================
def build_problem_summary(sensor_name, direction, normal_mean, anomaly_mean, count, max_consecutive):
    """
    담당자가 바로 이해할 수 있는 1줄 자연어 요약
    예) "정상(약 97.5℃)보다 1.3℃ 낮은 값이 14회 발생, 최대 25분 연속"
    """
    unit = extract_unit(sensor_name)
    diff = abs(anomaly_mean - normal_mean)

    # 소수점 자릿수: 차이가 매우 작으면 2자리, 아니면 1자리
    fmt = '.2f' if diff < 0.1 else '.1f'
    diff_str   = f"{diff:{fmt}}{unit}"
    normal_str = f"약 {normal_mean:{fmt}}{unit}" if unit else f"약 {normal_mean:.2f}"

    direction_text = '높은' if direction == '상승(▲)' else '낮은'
    summary = f"정상({normal_str})보다 {diff_str} {direction_text} 값이 {count}회 발생"
    if max_consecutive > 1:
        summary += f", 최대 {max_consecutive}분 연속"
    return summary


# ==========================================
# 3-C. 권고 조치 텍스트 생성
# ==========================================
def build_recommendation(sensor_name, risk, direction, count_ratio, max_consecutive, trend, max_z):
    """
    담당자가 바로 실행 가능한 번호 목록 형태 조치 사항 생성
    최대 4개 항목, \n 구분
    """
    parts = []
    name_upper = sensor_name.upper()

    # ① 긴급도
    if risk == '위험':
        parts.append('즉시 현장 확인 및 팀장 보고')
    elif risk == '주의':
        parts.append('점검 주기 단축 및 추이 집중 모니터링')
    else:
        parts.append('정기 점검 일정 준수')

    # ② 센서 종류별 점검 항목
    if 'TEMP' in name_upper or '온도' in sensor_name:
        if direction == '상승(▲)':
            parts.append('히터 과부하·냉각 불량·단열재 손상 여부 점검')
        else:
            parts.append('히터 단선·전원 공급 상태 및 열전대(TC) 동작 여부 점검')
    elif 'CV' in name_upper or '밸브' in sensor_name or 'VALVE' in name_upper:
        if direction == '상승(▲)':
            parts.append('액추에이터 과작동·포지셔너 오류 점검')
        else:
            parts.append('밸브 고착·공압 공급 불량 점검')
    elif 'PRESS' in name_upper or '압력' in sensor_name:
        if direction == '상승(▲)':
            parts.append('배관 막힘·안전밸브 작동 여부 확인')
        else:
            parts.append('배관 누설·펌프 성능 저하 확인')
    elif 'FLOW' in name_upper or '유량' in sensor_name:
        parts.append('배관 막힘·유량계 오염·펌프 성능 점검')
    elif 'SCR' in name_upper or '전류' in sensor_name:
        if direction == '상승(▲)':
            parts.append('SCR 과부하·접촉 불량 및 배선 상태 점검')
        else:
            parts.append('SCR 소자 불량·단선 및 전원 공급 상태 점검')
    elif '출력' in sensor_name:
        parts.append('제어 루프(PID) 출력 이상 여부 및 설정값 확인')

    # ③ 연속 이상
    if max_consecutive >= 30:
        parts.append(f'{max_consecutive}분 연속 이상 지속 → 설비 정지 후 정밀 점검 검토')
    elif max_consecutive >= 10:
        parts.append(f'{max_consecutive}분 연속 이상 → 긴급 점검 권고')

    # ④ 시계열 추세
    if trend == '상승추세(▲)':
        parts.append('시간 경과에 따른 상승 추세 → 마모·열화 진행 중, 수명 주기 점검')
    elif trend == '하강추세(▼)':
        parts.append('시간 경과에 따른 하강 추세 → 성능 저하 진행 중, 예방 정비 계획 수립')

    # 최대 4개 항목으로 절단
    parts = parts[:4]
    numbers = ['①', '②', '③', '④']
    return '\n'.join(f'{numbers[i]} {p}' for i, p in enumerate(parts))


# ==========================================
# 4. 핵심 분석 엔진 (고도화)
# ==========================================
def compute_enhanced_metrics(df, sensor_map):
    """
    Isolation Forest + Z-score + 이동평균 + 연속이상 + 추세분석을 통합한
    고도화 센서 이상 분석 엔진
    """
    print("[2/5] 고도화 이상 분석 수행 중 (Z-score·이동평균·연속이상·추세 포함)...")

    results = []
    anomaly_details = []

    clf = IsolationForest(contamination=0.01, random_state=42, n_jobs=-1)
    MA_WINDOW = 30  # 이동평균 윈도우 (데이터 포인트 기준)

    total = len(sensor_map)
    for idx_count, (col_idx, meta) in enumerate(sensor_map.items()):
        if (idx_count + 1) % 200 == 0:
            print(f"   ... {idx_count + 1}/{total} 센서 처리 완료")

        series = df[col_idx].reset_index(drop=True).astype(float)

        # 분석 제외 조건: 표준편차 0 또는 0/1 디지털 신호
        if series.std() < 1e-9:
            continue
        unique_vals = series.dropna().unique()
        if len(unique_vals) <= 2 and set(np.round(unique_vals)).issubset({0.0, 1.0}):
            continue
        # 유효 데이터 90% 미만 제외
        if series.isna().mean() > 0.1:
            continue

        # --- 이동평균(MA) 및 이탈량 계산 ---
        ma = series.rolling(window=MA_WINDOW, min_periods=1).mean()
        ma_std = series.rolling(window=MA_WINDOW, min_periods=1).std().fillna(0)
        ma_deviation = series - ma  # 이동평균 이탈량 (양수=상방 이탈, 음수=하방 이탈)

        # --- 전체 Z-score ---
        global_mean = series.mean()
        global_std = series.std() + 1e-9
        z_scores = (series - global_mean) / global_std

        # --- Isolation Forest 이상치 탐지 ---
        X = series.values.reshape(-1, 1)
        clf.fit(X)
        preds = clf.predict(X)
        if_scores = clf.decision_function(X)  # 음수일수록 이상치

        outlier_mask = preds == -1
        outlier_idx = np.where(outlier_mask)[0]
        count = len(outlier_idx)

        if count == 0:
            continue

        # --- 연속 이상 구간 탐지 ---
        max_consecutive = 0
        current_consecutive = 0
        consecutive_runs = []
        run_start = None

        for k in range(len(series)):
            if outlier_mask[k]:
                current_consecutive += 1
                if current_consecutive == 1:
                    run_start = k
                if current_consecutive > max_consecutive:
                    max_consecutive = current_consecutive
            else:
                if current_consecutive > 0:
                    consecutive_runs.append((run_start, k - 1, current_consecutive))
                current_consecutive = 0

        if current_consecutive > 0:
            consecutive_runs.append((run_start, len(series) - 1, current_consecutive))

        # --- 추세 분석 (선형 회귀 기울기) ---
        x_axis = np.arange(len(series))
        slope, intercept = np.polyfit(x_axis, series.values, 1)
        # 기울기를 평균값 대비 상대적 크기로 정규화
        rel_slope = slope / (abs(global_mean) + 1e-9)
        if rel_slope > 0.0001:
            trend = "상승추세(▲)"
        elif rel_slope < -0.0001:
            trend = "하강추세(▼)"
        else:
            trend = "안정(─)"

        # --- 이상치 통계 ---
        anomalies = series.iloc[outlier_idx]
        anomaly_mean = anomalies.mean()
        direction = "상승(▲)" if anomaly_mean > global_mean else "하강(▼)"
        change_ratio = (abs(anomaly_mean - global_mean) / (abs(global_mean) + 1e-9)) * 100

        z_at_anomalies = z_scores.iloc[outlier_idx]
        max_z = z_at_anomalies.abs().max()

        ma_dev_at_anomalies = ma_deviation.iloc[outlier_idx]
        max_ma_dev = ma_dev_at_anomalies.abs().max()
        avg_ma_dev = ma_dev_at_anomalies.abs().mean()

        count_ratio = count / len(series) * 100
        if_score = abs(if_scores[outlier_idx].mean())

        # 전반부/후반부 이상 분포
        n = len(series)
        early_count = int(np.sum(outlier_idx < n // 2))
        late_count = int(np.sum(outlier_idx >= n // 2))

        # --- 위험 등급 ---
        risk = classify_risk(if_score, max_z, max_consecutive, count_ratio, change_ratio)

        # --- 자연어 문제 요약 ---
        problem_summary = build_problem_summary(
            meta['name'], direction, global_mean, anomaly_mean, count, max_consecutive
        )

        # --- 번호 목록 조치 사항 ---
        rec = build_recommendation(
            meta['name'], risk, direction,
            count_ratio, max_consecutive, trend, max_z
        )

        # --- 처음/마지막 이상 발생 시각 ---
        times = df['Time'].reset_index(drop=True)
        first_anomaly = str(times.iloc[outlier_idx[0]]) if len(outlier_idx) > 0 else ''
        last_anomaly  = str(times.iloc[outlier_idx[-1]]) if len(outlier_idx) > 0 else ''

        results.append({
            'Col_Idx': col_idx,
            'iba Tag': meta['tag'],
            '명칭': meta['name'],
            '위험등급': risk,
            '문제 요약': problem_summary,
            '이상징후 건수': count,
            '이상 비율(%)': round(count_ratio, 2),
            '연속이상 최대구간(건)': max_consecutive,
            '처음 이상 발생': first_anomaly,
            '마지막 이상 발생': last_anomaly,
            '정상 평균': round(global_mean, 2),
            '이상구간 평균': round(anomaly_mean, 2),
            '권고 조치': rec,
            '패턴 방향': direction,
            '시계열 추세': trend,
            '정상 대비 변화율(%)': round(change_ratio, 2),
            '최대 Z-score': round(max_z, 2),
            '이동평균 최대이탈': round(max_ma_dev, 2),
            '이동평균 평균이탈': round(avg_ma_dev, 2),
            '전반부 이상건수': early_count,
            '후반부 이상건수': late_count,
            '이상치 최대값': round(anomalies.max(), 2),
            '이상치 최소값': round(anomalies.min(), 2),
            'IF 심각도 점수': round(if_score, 4),
            'AI 위험도 판단': '분석 대기'
        })

        # --- 이상치 상세 이력 (타임라인) — 담당자 친화형 ---
        unit = extract_unit(meta['name'])
        for oi in outlier_idx:
            measured = float(series.iloc[oi])
            diff_val  = measured - global_mean
            sign      = '+' if diff_val >= 0 else ''
            diff_text = f"{sign}{diff_val:.1f}{unit}"
            anomaly_details.append({
                'iba Tag': meta['tag'],
                '명칭': meta['name'],
                '위험등급': risk,
                '발생 시각': str(times.iloc[oi]) if oi < len(times) else '',
                '측정값': round(measured, 2),
                '정상 기준값': round(global_mean, 2),
                '정상 대비 차이': diff_text,
                '이상 방향': direction
            })

    print(f"   ... 분석 완료 ({len(results)}개 센서에서 이상 징후 감지)")

    if not results:
        return pd.DataFrame(), pd.DataFrame()

    df_results = pd.DataFrame(results).sort_values(by='IF 심각도 점수', ascending=False).reset_index(drop=True)
    df_details = pd.DataFrame(anomaly_details)

    return df_results, df_details


# ==========================================
# 5. 공통 스타일 정의
# ==========================================
def get_styles():
    styles = {
        'fill_danger':    PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid'),
        'fill_warning':   PatternFill(start_color="FFC000", end_color="FFC000", fill_type='solid'),
        'fill_ok':        PatternFill(start_color="92D050", end_color="92D050", fill_type='solid'),
        'fill_header':    PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type='solid'),
        'fill_subheader': PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type='solid'),
        'fill_row_alt':   PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type='solid'),
        'fill_row_light': PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type='solid'),
        'fill_danger_lt': PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type='solid'),
        'fill_warning_lt':PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type='solid'),
        'fill_ok_lt':     PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type='solid'),
        'font_header':    Font(color="FFFFFF", bold=True, size=11),
        'font_title':     Font(color="FFFFFF", bold=True, size=13),
        'font_danger':    Font(color="C00000", bold=True),
        'font_warning':   Font(color="7F4F00", bold=True),
        'font_ok':        Font(color="375623", bold=True),
        'font_bold':      Font(bold=True),
        'border_thin': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        ),
        'border_medium': Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'),  bottom=Side(style='medium')
        ),
        'align_center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'align_left':   Alignment(horizontal='left',   vertical='center', wrap_text=True),
        'align_top':    Alignment(horizontal='left',   vertical='top',    wrap_text=True),
    }
    return styles


def apply_risk_style(cell, risk, styles, light=False):
    """위험 등급에 따라 셀 색상·폰트 적용"""
    if risk == "위험":
        cell.fill = styles['fill_danger_lt'] if light else styles['fill_danger']
        cell.font = styles['font_danger'] if light else Font(color="FFFFFF", bold=True)
    elif risk == "주의":
        cell.fill = styles['fill_warning_lt'] if light else styles['fill_warning']
        cell.font = styles['font_warning'] if light else Font(color="7F4F00", bold=True)
    else:
        cell.fill = styles['fill_ok_lt'] if light else styles['fill_ok']
        cell.font = styles['font_ok'] if light else Font(color="375623", bold=True)


# ==========================================
# 6. Sheet 1: 종합 요약 대시보드
# ==========================================
def create_dashboard_sheet(wb, df_summary, styles):
    ws = wb.active
    ws.title = "종합 요약 대시보드"
    ws.sheet_view.showGridLines = False

    danger_count  = int((df_summary['위험등급'] == '위험').sum())
    warning_count = int((df_summary['위험등급'] == '주의').sum())
    ok_count      = int((df_summary['위험등급'] == '정상').sum())
    total_sensors = len(df_summary)

    # ── 타이틀 영역 ──────────────────────────────────────────────
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = "소성로 히터 센서 이상 분석 리포트  |  포스코 퓨처엠 설비 예지보전"
    title_cell.fill = styles['fill_header']
    title_cell.font = Font(color="FFFFFF", bold=True, size=15)
    title_cell.alignment = styles['align_center']
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:L2')
    ts_cell = ws['A2']
    ts_cell.value = f"분석 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}     |     분석 대상: {total_sensors:,}개 센서"
    ts_cell.fill = styles['fill_subheader']
    ts_cell.font = Font(color="FFFFFF", size=10)
    ts_cell.alignment = styles['align_center']
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10

    # ── 위험 등급 카드 영역 ───────────────────────────────────────
    def write_card(start_col, label, count, fill, font_color, desc):
        col_l = openpyxl.utils.get_column_letter
        # 라벨 셀
        ws.merge_cells(f'{col_l(start_col)}4:{col_l(start_col+3)}4')
        c = ws.cell(row=4, column=start_col, value=label)
        c.fill = fill; c.font = Font(color=font_color, bold=True, size=12)
        c.alignment = styles['align_center']
        ws.row_dimensions[4].height = 22

        ws.merge_cells(f'{col_l(start_col)}5:{col_l(start_col+3)}5')
        c = ws.cell(row=5, column=start_col, value=count)
        c.fill = fill; c.font = Font(color=font_color, bold=True, size=28)
        c.alignment = styles['align_center']
        ws.row_dimensions[5].height = 48

        ws.merge_cells(f'{col_l(start_col)}6:{col_l(start_col+3)}6')
        c = ws.cell(row=6, column=start_col, value=desc)
        c.fill = fill; c.font = Font(color=font_color, size=9)
        c.alignment = styles['align_center']
        ws.row_dimensions[6].height = 18

    write_card(1,  "⚠  위험  센서",  f"{danger_count}개",
               styles['fill_danger'],  "FFFFFF",
               "즉시 현장 점검 필요 | 설비 중단 검토")
    write_card(5,  "⚡  주의  센서",  f"{warning_count}개",
               styles['fill_warning'], "7F4F00",
               "점검 주기 단축 필요 | 추이 집중 모니터링")
    write_card(9,  "✓  정상  센서",  f"{ok_count}개",
               styles['fill_ok'],      "375623",
               "현 상태 유지 | 정기 점검 일정 준수")

    ws.row_dimensions[7].height = 12

    # ── 판정 기준 안내 (평문으로) ────────────────────────────────
    ws.merge_cells('A8:G8')
    c = ws.cell(row=8, column=1,
                value="【 판정 기준 】  "
                      "■ 위험: 이상이 전체의 5% 초과 또는 20분 이상 연속 발생  |  "
                      "■ 주의: 이상이 전체의 2% 초과 또는 10분 이상 연속 발생  |  "
                      "■ 정상: 위 기준 미달")
    c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type='solid')
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = styles['align_center']
    ws.row_dimensions[8].height = 16

    ws.row_dimensions[9].height = 10

    # ── TOP 20 위험 센서 테이블 (7개 컬럼) ──────────────────────
    headers_top20    = ["순위", "위험도", "센서 명칭", "무슨 문제?", "지금 해야 할 일", "이상 횟수", "최근 이상 시각"]
    col_widths_top20 = [6,      9,       40,          50,           55,               10,          22]

    ws.merge_cells('A10:G10')
    c = ws.cell(row=10, column=1, value="▶  TOP 20 위험 센서  (위험도 순 — 즉시 조치 대상부터 표시)")
    c.fill = styles['fill_subheader']
    c.font = Font(color="FFFFFF", bold=True, size=11)
    c.alignment = styles['align_center']
    ws.row_dimensions[10].height = 20

    for ci, (h, w) in enumerate(zip(headers_top20, col_widths_top20), 1):
        c = ws.cell(row=11, column=ci, value=h)
        c.fill = styles['fill_header']
        c.font = styles['font_header']
        c.alignment = styles['align_center']
        c.border = styles['border_thin']
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[11].height = 20

    top20    = df_summary.head(20)
    col_keys = [None, '위험등급', '명칭', '문제 요약', '권고 조치', '이상징후 건수', '마지막 이상 발생']

    for row_i, (_, row) in enumerate(top20.iterrows(), 12):
        risk_val = row['위험등급']
        row_fill = styles['fill_danger_lt'] if risk_val == '위험' else (
                   styles['fill_warning_lt'] if risk_val == '주의' else styles['fill_ok_lt'])

        for ci, key in enumerate(col_keys, 1):
            val = (row_i - 11) if key is None else row[key]
            c = ws.cell(row=row_i, column=ci, value=val)
            c.border = styles['border_thin']

            if ci == 2:
                apply_risk_style(c, risk_val, styles, light=False)
                c.alignment = styles['align_center']
            elif key in ('권고 조치', '문제 요약', '명칭'):
                c.alignment = styles['align_top']
                c.fill = row_fill
            elif ci == 1:
                c.alignment = styles['align_center']
                c.font = Font(bold=True)
                c.fill = row_fill
            else:
                c.alignment = styles['align_center']
                c.fill = row_fill

        ws.row_dimensions[row_i].height = 70  # 번호 목록 4줄 표시 확보


# ==========================================
# 7. Sheet 2: 센서별 상세 분석표
# ==========================================
def create_sensor_detail_sheet(wb, df_summary, styles):
    ws = wb.create_sheet(title="전체 센서 점검표")
    ws.sheet_view.showGridLines = False

    # 담당자 주요 컬럼 앞에, 기술 통계 뒤에 배치
    front_cols = [
        '위험등급', '명칭', '문제 요약', '권고 조치',
        '이상징후 건수', '연속이상 최대구간(건)',
        '처음 이상 발생', '마지막 이상 발생',
        '정상 평균', '이상구간 평균',
    ]
    tech_cols = [
        c for c in df_summary.columns
        if c not in front_cols and c != 'Col_Idx'
    ]
    display_cols = [c for c in front_cols if c in df_summary.columns] + \
                   [c for c in tech_cols if c in df_summary.columns]

    # 헤더 표시명 별칭
    header_alias = {
        '연속이상 최대구간(건)': '이상 지속\n(최대분)',
        '이상징후 건수':         '이상 횟수',
        '이상구간 평균':         '이상 평균',
        '처음 이상 발생':        '처음 발생',
        '마지막 이상 발생':      '마지막 발생',
        '권고 조치':             '조치 사항',
    }

    col_widths = {
        '위험등급': 9, '명칭': 38, '문제 요약': 50, '권고 조치': 55,
        '이상징후 건수': 11, '연속이상 최대구간(건)': 14,
        '처음 이상 발생': 22, '마지막 이상 발생': 22,
        '정상 평균': 12, '이상구간 평균': 12,
        'iba Tag': 14, '이상 비율(%)': 11, '최대 Z-score': 12,
        'IF 심각도 점수': 13, '이동평균 최대이탈': 16, '이동평균 평균이탈': 16,
        '패턴 방향': 11, '시계열 추세': 11, '정상 대비 변화율(%)': 14,
        '이상치 최대값': 12, '이상치 최소값': 12,
        '전반부 이상건수': 13, '후반부 이상건수': 13, 'AI 위험도 판단': 55,
    }

    # 타이틀
    last_col = openpyxl.utils.get_column_letter(len(display_cols))
    ws.merge_cells(f'A1:{last_col}1')
    c = ws.cell(row=1, column=1, value="전체 센서 점검표  ─  위험도 순 정렬  (왼쪽 10개 컬럼이 핵심 정보)")
    c.fill = styles['fill_header']
    c.font = styles['font_title']
    c.alignment = styles['align_center']
    ws.row_dimensions[1].height = 28

    # 헤더 행
    for ci, col in enumerate(display_cols, 1):
        header_text = header_alias.get(col, col)
        c = ws.cell(row=2, column=ci, value=header_text)
        c.fill = styles['fill_subheader']
        c.font = styles['font_header']
        c.alignment = styles['align_center']
        c.border = styles['border_thin']
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_widths.get(col, 13)
    ws.row_dimensions[2].height = 30

    # 데이터 행
    for row_i, (_, row) in enumerate(df_summary.iterrows(), 3):
        risk_val = row['위험등급']
        alt_fill = styles['fill_row_alt'] if row_i % 2 == 0 else styles['fill_row_light']

        for ci, col in enumerate(display_cols, 1):
            val = row[col] if col in row.index else ''
            c = ws.cell(row=row_i, column=ci, value=val)
            c.border = styles['border_thin']

            if col == '위험등급':
                apply_risk_style(c, risk_val, styles, light=False)
                c.alignment = styles['align_center']
            elif col in ('권고 조치', 'AI 위험도 판단', '명칭', '문제 요약'):
                c.alignment = styles['align_top']
                c.fill = alt_fill
            else:
                c.alignment = styles['align_center']
                c.fill = alt_fill

            # 연속이상 조건부 색상 (담당자가 한눈에 보도록)
            if col == '연속이상 최대구간(건)' and isinstance(val, (int, float)):
                if val > 20:
                    c.fill = styles['fill_danger_lt'];  c.font = styles['font_danger']
                elif val > 10:
                    c.fill = styles['fill_warning_lt']; c.font = styles['font_warning']

            if col == '이상 비율(%)' and isinstance(val, (int, float)):
                if val > 5:
                    c.fill = styles['fill_danger_lt'];  c.font = styles['font_danger']
                elif val > 2:
                    c.fill = styles['fill_warning_lt']; c.font = styles['font_warning']

        ws.row_dimensions[row_i].height = 75  # 번호 목록 4줄 충분히 표시


# ==========================================
# 8. Sheet 3: 이상치 상세 이력 (타임라인)
# ==========================================
def create_anomaly_timeline_sheet(wb, df_details, styles):
    if df_details.empty:
        return

    ws = wb.create_sheet(title="이상 발생 기록")
    ws.sheet_view.showGridLines = False

    # 위험 등급 → 발생 시각 순 정렬
    risk_order = {'위험': 0, '주의': 1, '정상': 2}
    df_sorted = df_details.copy()
    df_sorted['_sort'] = df_sorted['위험등급'].map(risk_order)
    df_sorted = df_sorted.sort_values(['_sort', '발생 시각']).drop(columns=['_sort'])

    # 담당자 친화형 컬럼만 표시 (Z-score·MA 제거)
    display_cols = ['위험등급', '발생 시각', '명칭', '측정값', '정상 기준값', '정상 대비 차이', '이상 방향', 'iba Tag']
    display_cols = [c for c in display_cols if c in df_sorted.columns]

    col_widths_map = {
        '위험등급': 9, '발생 시각': 22, '명칭': 38,
        '측정값': 12, '정상 기준값': 13, '정상 대비 차이': 14,
        '이상 방향': 11, 'iba Tag': 14,
    }

    ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(display_cols))}1')
    c = ws.cell(row=1, column=1,
                value=f"이상 발생 기록  ─  총 {len(df_sorted):,}건  (위험 등급 → 발생 시각 순)")
    c.fill = styles['fill_header']
    c.font = styles['font_title']
    c.alignment = styles['align_center']
    ws.row_dimensions[1].height = 28

    # 헤더
    for ci, col in enumerate(display_cols, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.fill = styles['fill_subheader']
        c.font = styles['font_header']
        c.alignment = styles['align_center']
        c.border = styles['border_thin']
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_widths_map.get(col, 13)
    ws.row_dimensions[2].height = 20

    # 데이터 (최대 5000건)
    df_limited = df_sorted.head(5000)
    for row_i, (_, row) in enumerate(df_limited.iterrows(), 3):
        risk_val = row['위험등급']
        alt_fill = styles['fill_row_alt'] if row_i % 2 == 0 else styles['fill_row_light']

        for ci, col in enumerate(display_cols, 1):
            val = row[col] if col in row.index else ''
            c = ws.cell(row=row_i, column=ci, value=val)
            c.border = styles['border_thin']

            if col == '위험등급':
                apply_risk_style(c, risk_val, styles, light=False)
                c.alignment = styles['align_center']
            elif col == '명칭':
                c.alignment = styles['align_left']
                c.fill = alt_fill
            else:
                c.alignment = styles['align_center']
                c.fill = alt_fill

        ws.row_dimensions[row_i].height = 16

    if len(df_sorted) > 5000:
        note_row = 2 + len(df_limited) + 1
        ws.cell(row=note_row, column=1,
                value=f"※ 전체 {len(df_sorted):,}건 중 상위 5,000건만 표시됩니다.")


# ==========================================
# 9. Sheet 4: 이동평균 분석 + 시계열 차트 (TOP N 센서)
# ==========================================
def create_moving_average_chart_sheet(wb, df, sensor_map, df_summary, styles, top_n=5):
    """
    위험도 상위 top_n 센서에 대해 원본값 vs 이동평균 시계열을 기록하고
    openpyxl LineChart로 시각화.
    """
    print(f"[4/5] 상위 {top_n}개 센서 시계열 차트 생성 중...")

    ws = wb.create_sheet(title="이동평균 분석(차트)")
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1,
                value=f"이동평균(MA30) 분석 ─ 위험도 상위 {top_n}개 센서 시계열 차트")
    c.fill = styles['fill_header']
    c.font = styles['font_title']
    c.alignment = styles['align_center']
    ws.row_dimensions[1].height = 28

    MA_WINDOW = 30
    top_sensors = df_summary.head(top_n)

    current_col = 1  # 데이터를 쓸 시작 열

    chart_anchors = []  # (차트 anchor 셀, 차트 객체) 리스트

    times = df['Time'].reset_index(drop=True)

    for sensor_rank, (_, sensor_row) in enumerate(top_sensors.iterrows()):
        col_idx = int(sensor_row['Col_Idx'])
        tag     = sensor_row['iba Tag']
        name    = sensor_row['명칭']
        risk    = sensor_row['위험등급']

        series = df[col_idx].reset_index(drop=True).astype(float)
        ma     = series.rolling(window=MA_WINDOW, min_periods=1).mean()

        # 이 센서의 데이터를 쓸 열 범위: (시각, 원본값, MA)
        col_time = current_col
        col_val  = current_col + 1
        col_ma   = current_col + 2

        # 섹션 제목
        header_row = 3 + sensor_rank * (len(series) + 6)

        ws.merge_cells(
            start_row=header_row, start_column=col_time,
            end_row=header_row,   end_column=col_ma
        )
        c = ws.cell(row=header_row, column=col_time,
                    value=f"[{sensor_rank+1}위] {tag}  |  {name}  |  위험등급: {risk}")
        c.fill = styles['fill_subheader']
        c.font = styles['font_header']
        c.alignment = styles['align_center']
        ws.row_dimensions[header_row].height = 22

        # 컬럼 헤더
        for ci, label in enumerate(['발생 시각', '원본값', f'이동평균(MA{MA_WINDOW})'], col_time):
            c = ws.cell(row=header_row + 1, column=ci, value=label)
            c.fill = styles['fill_header']
            c.font = styles['font_header']
            c.alignment = styles['align_center']
            c.border = styles['border_thin']
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22
        ws.row_dimensions[header_row + 1].height = 18

        # 데이터 쓰기
        data_start_row = header_row + 2
        for i, (t, v, m) in enumerate(zip(times, series, ma)):
            r = data_start_row + i
            ws.cell(row=r, column=col_time, value=str(t)).alignment = styles['align_center']
            ws.cell(row=r, column=col_val,  value=round(float(v), 2)).alignment = styles['align_center']
            ws.cell(row=r, column=col_ma,   value=round(float(m), 2)).alignment = styles['align_center']
            ws.row_dimensions[r].height = 14

        data_end_row = data_start_row + len(series) - 1

        # LineChart 생성
        chart = LineChart()
        chart.title      = f"{tag} │ {name[:30]}  (위험등급: {risk})"
        chart.style      = 10
        chart.y_axis.title = "센서값"
        chart.x_axis.title = "시간 인덱스"
        chart.width  = 22
        chart.height = 12

        # 원본값 시리즈
        ref_val = Reference(ws, min_col=col_val, min_row=data_start_row,
                            max_row=data_end_row)
        chart.add_data(ref_val, titles_from_data=False)
        chart.series[0].title = SeriesLabel(v="원본값")
        chart.series[0].graphicalProperties.line.solidFill = "FF0000" if risk == "위험" else (
            "FFC000" if risk == "주의" else "4472C4")
        chart.series[0].graphicalProperties.line.width = 15000  # 1.5pt

        # 이동평균 시리즈
        ref_ma = Reference(ws, min_col=col_ma, min_row=data_start_row,
                           max_row=data_end_row)
        chart.add_data(ref_ma, titles_from_data=False)
        chart.series[1].title = SeriesLabel(v=f"이동평균(MA{MA_WINDOW})")
        chart.series[1].graphicalProperties.line.solidFill = "595959"
        chart.series[1].graphicalProperties.line.width = 20000  # 2pt (굵게)

        # 차트 삽입 위치 (데이터 옆 열)
        anchor_col = col_ma + 2
        anchor_cell = f"{openpyxl.utils.get_column_letter(anchor_col)}{header_row}"
        chart_anchors.append((anchor_cell, chart))

    # 차트 워크시트에 삽입
    for anchor, chart in chart_anchors:
        ws.add_chart(chart, anchor)

    # 열 너비 조정
    for ci in range(1, current_col + 1):
        col_letter = openpyxl.utils.get_column_letter(ci)
        if ws.column_dimensions[col_letter].width < 10:
            ws.column_dimensions[col_letter].width = 14


# ==========================================
# 10. Sheet 5: 구역별 현황 요약 (담당자 친화형)
# ==========================================
def create_zone_summary_sheet(wb, df_summary, styles):
    """
    센서명 prefix(예비소성/본소성A/본소성B)로 구역 분류 후
    구역별 위험/주의/정상 현황과 대표 위험 센서를 한눈에 표시.
    """
    ws = wb.create_sheet(title="구역별 현황")
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1,
                value="구역별 현황 요약  ─  설비 구역별 이상 센서 현황 및 대표 위험 센서")
    c.fill = styles['fill_header']
    c.font = styles['font_title']
    c.alignment = styles['align_center']
    ws.row_dimensions[1].height = 28

    # 구역 분류 (실제 센서명 prefix 기준)
    def get_zone(name):
        if '#01_예비소성' in name:
            return '예비소성 구역'
        elif '#01_A_' in name:
            return '본소성 A 구역'
        elif '#01_B_' in name:
            return '본소성 B 구역'
        else:
            return '기타'

    zone_data = {}
    for _, row in df_summary.iterrows():
        zone = get_zone(row['명칭'])
        zone_data.setdefault(zone, []).append(row)

    zone_order = ['예비소성 구역', '본소성 A 구역', '본소성 B 구역', '기타']

    # 구역 카드 헤더
    headers_z = ["구역명", "전체 센서", "위험", "주의", "정상", "대표 위험 센서명", "주요 문제 요약"]
    col_widths_z = [18, 11, 9, 9, 9, 42, 55]

    for ci, (h, w) in enumerate(zip(headers_z, col_widths_z), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = styles['fill_subheader']
        c.font = styles['font_header']
        c.alignment = styles['align_center']
        c.border = styles['border_thin']
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    # 구역별 집계 출력
    for row_i, zone in enumerate(zone_order, 3):
        rows = zone_data.get(zone, [])
        total   = len(rows)
        danger  = sum(1 for r in rows if r['위험등급'] == '위험')
        warning = sum(1 for r in rows if r['위험등급'] == '주의')
        ok      = sum(1 for r in rows if r['위험등급'] == '정상')

        # 대표 위험 센서: 위험 > 주의 > 정상 순, IF 심각도 높은 것
        sorted_rows = sorted(rows,
            key=lambda r: (0 if r['위험등급'] == '위험' else 1 if r['위험등급'] == '주의' else 2,
                           -r['IF 심각도 점수']))
        rep = sorted_rows[0] if sorted_rows else None
        rep_name    = rep['명칭'][:40] if rep is not None else '-'
        rep_summary = rep['문제 요약'] if rep is not None else '-'
        rep_risk    = rep['위험등급'] if rep is not None else '정상'

        # 구역 전체 위험도
        zone_risk = '위험' if danger > 0 else ('주의' if warning > 0 else '정상')

        # 배경색
        zone_fill = (styles['fill_danger_lt'] if zone_risk == '위험' else
                     styles['fill_warning_lt'] if zone_risk == '주의' else
                     styles['fill_ok_lt'])

        row_vals = [zone, total, danger, warning, ok, rep_name, rep_summary]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_i, column=ci, value=val)
            c.border = styles['border_thin']
            if ci == 3 and danger > 0:
                c.fill = styles['fill_danger_lt']; c.font = styles['font_danger']
                c.alignment = styles['align_center']
            elif ci == 4 and warning > 0:
                c.fill = styles['fill_warning_lt']; c.font = styles['font_warning']
                c.alignment = styles['align_center']
            elif ci == 5 and ok > 0:
                c.fill = styles['fill_ok_lt']; c.font = styles['font_ok']
                c.alignment = styles['align_center']
            elif ci in (6, 7):
                c.alignment = styles['align_top']
                c.fill = zone_fill
            elif ci == 1:
                c.fill = zone_fill
                c.font = Font(bold=True)
                c.alignment = styles['align_center']
            else:
                c.fill = zone_fill
                c.alignment = styles['align_center']
        ws.row_dimensions[row_i].height = 50

    ws.row_dimensions[7].height = 14

    # ── 구역별 위험 센서 상세 테이블 ─────────────────────────────
    detail_row = 8
    for zone in zone_order:
        rows = zone_data.get(zone, [])
        danger_rows = [r for r in rows if r['위험등급'] == '위험']
        warning_rows = [r for r in rows if r['위험등급'] == '주의']
        show_rows = (sorted(danger_rows, key=lambda r: -r['IF 심각도 점수']) +
                     sorted(warning_rows, key=lambda r: -r['IF 심각도 점수']))[:10]

        if not show_rows:
            continue

        # 구역 섹션 헤더
        ws.merge_cells(f'A{detail_row}:G{detail_row}')
        c = ws.cell(row=detail_row, column=1,
                    value=f"[ {zone} ] — 위험·주의 센서 상세 (최대 10개)")
        c.fill = styles['fill_subheader']
        c.font = styles['font_header']
        c.alignment = styles['align_center']
        ws.row_dimensions[detail_row].height = 18
        detail_row += 1

        sub_headers = ["위험도", "센서 명칭", "무슨 문제?", "지금 해야 할 일", "이상 횟수", "처음 발생", "최근 발생"]
        sub_widths  = [9, 42, 50, 55, 10, 22, 22]
        for ci, (h, w) in enumerate(zip(sub_headers, sub_widths), 1):
            c = ws.cell(row=detail_row, column=ci, value=h)
            c.fill = styles['fill_header']
            c.font = styles['font_header']
            c.alignment = styles['align_center']
            c.border = styles['border_thin']
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width, w)
        ws.row_dimensions[detail_row].height = 18
        detail_row += 1

        for r in show_rows:
            risk_val = r['위험등급']
            row_fill = (styles['fill_danger_lt'] if risk_val == '위험' else styles['fill_warning_lt'])
            vals = [risk_val, r['명칭'], r['문제 요약'], r['권고 조치'],
                    r['이상징후 건수'], r['처음 이상 발생'], r['마지막 이상 발생']]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=detail_row, column=ci, value=val)
                c.border = styles['border_thin']
                if ci == 1:
                    apply_risk_style(c, risk_val, styles, light=False)
                    c.alignment = styles['align_center']
                elif ci in (3, 4):
                    c.alignment = styles['align_top']
                    c.fill = row_fill
                else:
                    c.alignment = styles['align_center']
                    c.fill = row_fill
            ws.row_dimensions[detail_row].height = 60
            detail_row += 1

        ws.row_dimensions[detail_row].height = 8
        detail_row += 1


# ==========================================
# 11. Sheet 6: 메트릭 정의 (고도화)
# ==========================================
def create_metrics_definition_sheet(wb, styles):
    """판정 기준 안내 — 기술 용어 없이 담당자 언어로 작성"""
    ws = wb.create_sheet(title="판정 기준 안내")
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value="판정 기준 안내  ─  이 리포트의 지표를 쉽게 이해하기 위한 참조표")
    c.fill = styles['fill_header']
    c.font = styles['font_title']
    c.alignment = styles['align_center']
    ws.row_dimensions[1].height = 28

    definition_data = [
        ["항목", "무슨 뜻인가?", "위험 기준", "주의 기준", "현장 조치"],
        [
            "이상 횟수",
            "측정 기간(약 24시간) 동안 정상 범위를 벗어난 측정값의 횟수.\n"
            "많을수록 설비 상태가 불안정한 것.",
            "전체의 5% 초과\n(1,440분 기준 약 72분 이상)",
            "전체의 2% 초과\n(약 29분 이상)",
            "이상 횟수 상위 센서부터 현장 확인 우선 실시."
        ],
        [
            "이상 지속\n(최대분)",
            "이상 상태가 끊김 없이 연속으로 이어진 가장 긴 시간(분).\n"
            "길수록 일시적인 전기 노이즈가 아닌 실제 설비 결함 가능성이 높음.",
            "20분 초과 연속\n→ 설비 정지 후 점검 검토",
            "10분 초과 연속\n→ 추이 집중 관찰",
            "연속 20분 초과: 운전 정지 후 정밀 점검.\n"
            "연속 10~20분: 점검 주기 단축."
        ],
        [
            "위험 / 주의 / 정상",
            "이상 횟수·연속 시간 등을 종합하여 자동 판정한 경보 등급.\n"
            "설비 담당자가 가장 먼저 확인해야 할 핵심 지표.",
            "'위험' = 즉시 현장 확인 필요",
            "'주의' = 점검 주기 단축 권고",
            "시트 1(대시보드) 빨간 카드에 표시된 위험 센서부터 처리."
        ],
        [
            "정상 기준값",
            "해당 센서의 24시간 전체 측정 평균값.\n"
            "정상 운전 상태의 대표 기준으로 사용.",
            "—",
            "—",
            "측정값이 이 기준과 크게 다를수록 이상 가능성 높음.\n"
            "'무슨 문제?' 컬럼에 자동 계산된 차이가 표시됨."
        ],
        [
            "정상 대비 차이",
            "각 이상 발생 시점의 실제 측정값이 정상 기준값보다\n"
            "얼마나 높거나(+) 낮은지(-) 보여주는 값.\n"
            "예: +2.3℃ → 정상보다 2.3℃ 높음",
            "차이가 클수록 심각",
            "—",
            "'이상 발생 기록' 시트에서 각 건별로 확인 가능.\n"
            "차이가 반복적으로 일정 방향이면 설비 점검 필요."
        ],
    ]

    col_widths_def = [18, 45, 28, 28, 45]
    fill_row_colors = ["DEEAF1", "F2F7FB"]

    for ci, w in enumerate(col_widths_def, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    for row_idx, row_data in enumerate(definition_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = styles['border_thin']
            c.alignment = Alignment(wrap_text=True, vertical='top')

            if row_idx == 2:
                c.fill = styles['fill_subheader']
                c.font = styles['font_header']
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                fill_color = fill_row_colors[(row_idx - 3) % 2]
                c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                if col_idx == 1:
                    c.font = Font(bold=True)

    ws.row_dimensions[2].height = 20
    for i in range(3, len(definition_data) + 2):
        ws.row_dimensions[i].height = 72


# ==========================================
# 12. AI 진단 실행 (상위 10개 센서)
# ==========================================
def run_ai_diagnosis(df_summary):
    print("[3/5] AI(GPT) 정밀 진단 수행 중 (상위 10개 센서)...")
    targets = df_summary.head(10)

    for idx in targets.index:
        row = df_summary.loc[idx]
        print(f"   > AI 분석: {row['명칭'][:40]} (위험등급: {row['위험등급']})")

        prompt = f"""
[역할]
당신은 포스코 퓨처엠 소성로 설비 예지보전 전문가입니다.
현장 엔지니어가 엑셀에서 바로 읽고 즉각 조치할 수 있도록 작성하십시오.

[분석 데이터]
- 센서 명칭: {row['명칭']}
- 위험 등급: {row['위험등급']}
- 전체 이상치 건수: {row['이상징후 건수']}건 (전체 대비 {row['이상 비율(%)']}%)
- 정상 평균: {row['정상 평균']}  /  이상구간 평균: {row['이상구간 평균']}
- 정상 대비 변화율: {row['정상 대비 변화율(%)']}%
- 최대 Z-score: {row['최대 Z-score']}σ
- 이동평균 최대이탈: {row['이동평균 최대이탈']}
- 연속 이상 최대: {row['연속이상 최대구간(건)']}건
- 패턴 방향: {row['패턴 방향']}  /  시계열 추세: {row['시계열 추세']}
- 전반부 이상: {row['전반부 이상건수']}건  /  후반부 이상: {row['후반부 이상건수']}건

[출력 규칙 – 반드시 준수]
1. 아래 형식만 사용, 항목당 1~2문장 이내
2. 고장 위험도는 반드시 % 숫자로 명시
3. 서론·결론·불필요한 설명 추가 금지

[출력 형식]
① 이상 분석: [관측된 이상 패턴 핵심 요약]
② 근거: [수치 근거 – 정상 대비 편차, Z-score, 연속이상 등]
③ 추정 원인: [히터 단선·과부하·센서 오류 등 구체적 원인 추정]
④ 고장 위험도: [XX%] – [위험 등급 근거 요약]
⑤ 즉시 조치: [현장 담당자가 바로 실행할 수 있는 구체적 행동]
"""
        ai_result = call_corporate_ai_api(prompt)
        df_summary.loc[idx, 'AI 위험도 판단'] = ai_result

    return df_summary


# ==========================================
# 13. 최종 엑셀 리포트 생성
# ==========================================
def create_smart_report(df, sensor_map, df_summary, df_details, output_path):
    print(f"[5/5] 엑셀 리포트 저장 중: {output_path}")

    # AI 진단 (API key 있을 때만)
    if POSCO_GPT_KEY:
        df_summary = run_ai_diagnosis(df_summary)

    styles = get_styles()
    wb = openpyxl.Workbook()

    # Sheet 1: 종합 요약 대시보드
    create_dashboard_sheet(wb, df_summary, styles)

    # Sheet 2: 센서별 상세 분석표
    create_sensor_detail_sheet(wb, df_summary, styles)

    # Sheet 3: 이상치 상세 이력
    create_anomaly_timeline_sheet(wb, df_details, styles)

    # Sheet 4: 이동평균 분석 + 차트
    create_moving_average_chart_sheet(wb, df, sensor_map, df_summary, styles, top_n=5)

    # Sheet 5: 구역별 현황 (기존 패턴분석 대체)
    create_zone_summary_sheet(wb, df_summary, styles)

    # Sheet 6: 판정 기준 안내 (단순화)
    create_metrics_definition_sheet(wb, styles)

    wb.save(output_path)
    print(f"   >> 저장 완료: {output_path}")


# ==========================================
# 14. 메인 실행 함수 (GUI)
# ==========================================
def main():
    print("=== [POSCO Future M] AI 예지보전 시스템 가동 ===")

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    if not POSCO_GPT_KEY:
        messagebox.showwarning(
            "키 확인",
            "환경변수 POSCO_GPT_KEY가 설정되지 않았습니다.\n"
            "AI 진단 기능은 비활성화되며, 나머지 분석은 정상 수행됩니다."
        )

    while True:
        file_path = filedialog.askopenfilename(
            parent=root,
            title="분석할 센서 데이터 파일(CSV/TXT) 선택",
            filetypes=[("CSV/TXT Files", "*.csv *.txt"), ("All Files", "*.*")]
        )

        if not file_path:
            print(">> 종료합니다.")
            break

        try:
            df, sensor_map = load_data(file_path)

            # 분석 전 유효 센서 사전 필터링 (전체 0 또는 상수 센서 제외 → 성능 최적화)
            print("[전처리] 유효 센서 필터링 중...")
            import numpy as _np
            valid_sensor_map = {}
            for _ci, _meta in sensor_map.items():
                _s = df[_ci].astype(float)
                if _s.std() < 1e-9:
                    continue
                _uv = _s.dropna().unique()
                if len(_uv) <= 2 and set(_np.round(_uv)).issubset({0.0, 1.0}):
                    continue
                valid_sensor_map[_ci] = _meta
            print(f"   >> 전체 {len(sensor_map)}개 센서 중 유효 {len(valid_sensor_map)}개 선별")

            df_result, df_details = compute_enhanced_metrics(df, valid_sensor_map)

            if df_result.empty:
                messagebox.showinfo("결과", "분석 가능한 아날로그 신호에서 특이점이 발견되지 않았습니다.")
            else:
                danger_n  = int((df_result['위험등급'] == '위험').sum())
                warning_n = int((df_result['위험등급'] == '주의').sum())
                ok_n      = int((df_result['위험등급'] == '정상').sum())

                timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
                base_name  = os.path.basename(file_path).rsplit('.', 1)[0]
                output_file = f"SmartReport_{base_name}_{timestamp}.xlsx"

                create_smart_report(df, sensor_map, df_result, df_details, output_file)

                msg = (
                    f"분석 완료!\n\n"
                    f"  위험  센서: {danger_n}개\n"
                    f"  주의  센서: {warning_n}개\n"
                    f"  정상  센서: {ok_n}개\n\n"
                    f"저장 위치: {output_file}\n\n"
                    f"계속 분석하시겠습니까?"
                )
                if not messagebox.askyesno("완료", msg, parent=root):
                    break

        except Exception as e:
            import traceback
            err_detail = traceback.format_exc()
            print(f"[Error]\n{err_detail}")
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다:\n{str(e)}", parent=root)

    root.destroy()


if __name__ == "__main__":
    main()
